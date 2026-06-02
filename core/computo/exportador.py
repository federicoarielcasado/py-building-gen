"""Exportador de presupuesto a Excel (.xlsx) y PDF.

Excel: una hoja por rubro + resumen ejecutivo + curva de inversión.
PDF:   documento de presentación con tabla de rubros y resumen.
"""

from __future__ import annotations

from pathlib import Path
from datetime import date
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from core.computo.analisis_precios import PresupuestoCompleto, RubroAnalizado

if TYPE_CHECKING:
    pass

# ---------------------------------------------------------------------------
# Colores y estilos (Excel)
# ---------------------------------------------------------------------------

_AZUL_HEADER  = "1A4F7A"
_GRIS_FILA    = "F0F4F8"
_AMARILLO_SUB = "FFF3CD"
_VERDE_TOTAL  = "D4EDDA"

_THIN = Side(style="thin", color="CCCCCC")
_BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_NUM_ARS = '#,##0'
_NUM_ARS_DEC = '#,##0.00'


def _header_cell(ws, row: int, col: int, value: str, *, bold: bool = True) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=bold, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=_AZUL_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _BORDER_THIN


def _data_cell(ws, row: int, col: int, value, *, numero: bool = False, negrita: bool = False, bg: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=9, bold=negrita)
    cell.alignment = Alignment(horizontal="right" if numero else "left", vertical="center")
    cell.border = _BORDER_THIN
    if numero and isinstance(value, (int, float)):
        cell.number_format = _NUM_ARS
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def exportar_excel(
    presupuesto: PresupuestoCompleto,
    path: Path | str,
    moneda: str = "ARS",
) -> Path:
    """Exporta el presupuesto a un archivo Excel con una hoja por rubro.

    Args:
        presupuesto: Presupuesto analizado.
        path: Ruta de salida del archivo .xlsx.
        moneda: "ARS" o "USD".

    Returns:
        Path al archivo generado.
    """
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # Hoja resumen ejecutivo
    _hoja_resumen(wb, presupuesto, moneda)

    # Una hoja por rubro
    for rubro in presupuesto.rubros:
        _hoja_rubro(wb, rubro, moneda)

    # Hoja curva de inversión (distribución mensual estimada)
    _hoja_curva(wb, presupuesto, moneda)

    wb.save(out)
    return out


def _hoja_resumen(wb: openpyxl.Workbook, ppto: PresupuestoCompleto, moneda: str) -> None:
    ws = wb.active
    ws.title = "RESUMEN EJECUTIVO"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14

    # Encabezado
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "py-building-gen — CÓMPUTO Y PRESUPUESTO"
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=_AZUL_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Fecha: {date.today().isoformat()}  |  Moneda: {moneda}  |  Precios: UOCRA Zona A + ICC INDEC + Sismat 2026"
    ws["A2"].font = Font(name="Calibri", size=9, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Cabecera tabla
    row = 4
    for col, titulo in enumerate(["N°", "Rubro", f"Subtotal ({moneda})", "% s/total"], 1):
        _header_cell(ws, row, col, titulo)
    ws.row_dimensions[row].height = 20

    row = 5
    total = ppto.total or 1.0
    for i, info in enumerate(ppto.resumen_por_rubro):
        bg = _GRIS_FILA if i % 2 == 0 else None
        _data_cell(ws, row, 1, info["numero"], bg=bg)
        _data_cell(ws, row, 2, info["nombre"], bg=bg)
        _data_cell(ws, row, 3, round(info["subtotal"]), numero=True, bg=bg)
        _data_cell(ws, row, 4, f"{info['pct_sobre_total']:.1f}%", bg=bg)
        row += 1

    # Líneas de totales
    def total_row(label: str, valor: float, bg: str) -> None:
        nonlocal row
        ws.merge_cells(f"A{row}:B{row}")
        _data_cell(ws, row, 1, label, negrita=True, bg=bg)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        _data_cell(ws, row, 3, round(valor), numero=True, negrita=True, bg=bg)
        row += 1

    row += 1
    total_row("COSTO DIRECTO", ppto.costo_directo, _AMARILLO_SUB)
    if ppto.incluir_gastos:
        total_row(f"Gastos generales ({ppto.pct_gastos_generales*100:.0f}%)", ppto.gastos_generales, _AMARILLO_SUB)
    if ppto.incluir_honorarios:
        total_row(f"Honorarios profesionales ({ppto.pct_honorarios*100:.0f}%)", ppto.honorarios, _AMARILLO_SUB)
    total_row("TOTAL PRESUPUESTO", ppto.total, _VERDE_TOTAL)


def _hoja_rubro(wb: openpyxl.Workbook, rubro: RubroAnalizado, moneda: str) -> None:
    nombre_hoja = f"{rubro.numero:02d} {rubro.nombre}"[:31]
    ws = wb.create_sheet(title=nombre_hoja)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    # Encabezado
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"RUBRO {rubro.numero} — {rubro.nombre.upper()}"
    c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=_AZUL_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    row = 3
    for col, titulo in enumerate(["Código", "Descripción", "Unidad", "Cantidad", f"P.U. ({moneda})", f"Subtotal ({moneda})"], 1):
        _header_cell(ws, row, col, titulo)

    row = 4
    for i, item in enumerate(rubro.items):
        bg = _GRIS_FILA if i % 2 == 0 else None
        _data_cell(ws, row, 1, item.codigo, bg=bg)
        _data_cell(ws, row, 2, item.descripcion, bg=bg)
        _data_cell(ws, row, 3, item.unidad, bg=bg)
        _data_cell(ws, row, 4, round(item.cantidad_obra, 2), numero=True, bg=bg)
        _data_cell(ws, row, 5, round(item.precio_unitario), numero=True, bg=bg)
        _data_cell(ws, row, 6, round(item.subtotal_obra), numero=True, bg=bg)
        row += 1

    # Subtotal del rubro
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    _data_cell(ws, row, 1, f"SUBTOTAL RUBRO {rubro.numero}", negrita=True, bg=_AMARILLO_SUB)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    _data_cell(ws, row, 6, round(rubro.subtotal), numero=True, negrita=True, bg=_AMARILLO_SUB)


def _hoja_curva(wb: openpyxl.Workbook, ppto: PresupuestoCompleto, moneda: str) -> None:
    """Distribución mensual estimada del gasto (curva de inversión simplificada)."""
    ws = wb.create_sheet(title="CURVA DE INVERSIÓN")

    # Distribución porcentual típica para edificio multifamiliar (18 meses)
    dist_pct = [2, 4, 6, 8, 9, 9, 8, 8, 7, 7, 6, 6, 5, 5, 4, 3, 2, 1]
    meses = len(dist_pct)
    total = ppto.total

    ws["A1"] = "Mes"
    ws["B1"] = f"Inversión mensual ({moneda})"
    ws["C1"] = f"Inversión acumulada ({moneda})"
    for col_idx, titulo in enumerate(["A", "B", "C"], 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)

    acum = 0.0
    for i, pct in enumerate(dist_pct, 1):
        mensual = total * pct / 100
        acum += mensual
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=round(mensual))
        ws.cell(row=i + 1, column=3, value=round(acum))

    # Gráfico de barras
    chart = BarChart()
    chart.type = "col"
    chart.title = "Curva de inversión"
    chart.y_axis.title = f"ARS"
    chart.x_axis.title = "Mes"
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=meses + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=meses + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    ws.add_chart(chart, "E2")


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def exportar_pdf(
    presupuesto: PresupuestoCompleto,
    path: Path | str,
    moneda: str = "ARS",
    nombre_proyecto: str = "Edificio Residencial — Caballito, CABA",
    autor: str = "Federico A. Casado",
) -> Path:
    """Exporta el presupuesto a un PDF de presentación.

    Args:
        presupuesto: Presupuesto analizado.
        path: Ruta de salida del archivo .pdf.
        moneda: "ARS" o "USD".
        nombre_proyecto: Nombre del proyecto.
        autor: Nombre del autor.

    Returns:
        Path al archivo generado.
    """
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(out),
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleCustom", parent=styles["Title"],
        fontSize=16, textColor=colors.HexColor("#1A4F7A"),
        spaceAfter=6,
    )
    h2_style = ParagraphStyle(
        "H2Custom", parent=styles["Heading2"],
        fontSize=11, textColor=colors.HexColor("#1A4F7A"),
        spaceBefore=12, spaceAfter=4,
    )
    normal = styles["Normal"]
    normal.fontSize = 9

    story = []

    # Título
    story.append(Paragraph("py-building-gen", title_style))
    story.append(Paragraph(f"Cómputo y Presupuesto — {nombre_proyecto}", h2_style))
    story.append(Paragraph(
        f"Autor: {autor}  |  Fecha: {date.today().isoformat()}  |  Moneda: {moneda}",
        normal,
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1A4F7A")))
    story.append(Spacer(1, 8 * mm))

    # Tabla resumen
    story.append(Paragraph("Resumen por rubro", h2_style))

    data_tabla = [["N°", "Rubro", f"Subtotal ({moneda})", "% s/total"]]
    for info in presupuesto.resumen_por_rubro:
        data_tabla.append([
            str(info["numero"]),
            info["nombre"],
            f"{round(info['subtotal']):,}",
            f"{info['pct_sobre_total']:.1f}%",
        ])
    data_tabla.append(["", "COSTO DIRECTO", f"{round(presupuesto.costo_directo):,}", ""])
    if presupuesto.incluir_gastos:
        data_tabla.append(["", "Gastos generales", f"{round(presupuesto.gastos_generales):,}", ""])
    if presupuesto.incluir_honorarios:
        data_tabla.append(["", "Honorarios profesionales", f"{round(presupuesto.honorarios):,}", ""])
    data_tabla.append(["", "TOTAL PRESUPUESTO", f"{round(presupuesto.total):,}", ""])

    col_widths = [15 * mm, 90 * mm, 40 * mm, 25 * mm]
    tabla = Table(data_tabla, colWidths=col_widths)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A4F7A")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("ALIGN",      (2, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -5), [colors.white, colors.HexColor("#F0F4F8")]),
        ("BACKGROUND", (0, -4), (-1, -4), colors.HexColor("#FFF3CD")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D4EDDA")),
        ("FONTNAME",   (0, -4), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 1), (-1, -1), 8),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tabla)
    story.append(Spacer(1, 8 * mm))

    # Notas
    story.append(Paragraph("Notas", h2_style))
    for nota in [
        "Precios de referencia: UOCRA Zona A + ICC INDEC + Sismat (mayo 2026).",
        "Los precios no incluyen IVA. Actualizar según índice CAC mensual.",
        "Las cantidades son estimadas a partir de parámetros volumétricos del edificio.",
        "Para proyecto ejecutivo, realizar cómputo detallado sobre planos aprobados.",
        "Generado automáticamente por py-building-gen — Federico A. Casado.",
    ]:
        story.append(Paragraph(f"• {nota}", normal))

    doc.build(story)
    return out
